from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
import random
import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages
from .models import (
    Mushroom, Quiz, QuizQuestion, QuizAnswer, QuizResult,
    Characteristic, CharacteristicOption, MushroomCharacteristic, Lookalike, UserAnswer
)

def home(request):
    """Главная страница"""
    return render(request, 'home.html')

def edible_mushrooms(request):
    """Съедобные грибы"""
    mushrooms = Mushroom.objects.filter(
        edibility__in=['edible', 'conditionally_edible']
    )
    
    tubular = mushrooms.filter(mushroom_type='tubular')
    lamellar = mushrooms.filter(mushroom_type='lamellar')
    other = mushrooms.filter(mushroom_type='other')
    
    return render(request, 'edible_mushrooms.html', {
        'tubular_mushrooms': tubular,
        'lamellar_mushrooms': lamellar,
        'other_mushrooms': other,
    })

def poisonous_mushrooms(request):
    """Ядовитые грибы"""
    mushrooms = Mushroom.objects.filter(
        edibility__in=['poisonous', 'deadly']
    )
    return render(request, 'poisonous_mushrooms.html', {
        'mushrooms': mushrooms
    })

def gallery(request):
    """Галерея всех грибов"""
    mushrooms = Mushroom.objects.all()
    return render(request, 'gallery.html', {
        'mushrooms': mushrooms
    })

def interactive_identifier(request):
    """Интерактивный определитель грибов"""
    characteristics = Characteristic.objects.prefetch_related('characteristicoption_set').all()
    
    if request.method == 'POST':
        selected_options = {}
        for key, value in request.POST.items():
            if key.startswith('char_') and value:  # Только если значение не пустое
                char_id = key.replace('char_', '')
                selected_options[char_id] = value
        
        # Всегда находим грибы, даже если не выбрано ни одной характеристики
        matching_mushrooms = find_matching_mushrooms(selected_options)
        
        results = []
        for mushroom in matching_mushrooms:
            lookalikes = Lookalike.objects.filter(mushroom=mushroom).select_related('lookalike')
            match_percentage = calculate_match_percentage(mushroom, selected_options)
            
            results.append({
                'mushroom': mushroom,
                'lookalikes': lookalikes,
                'match_percentage': match_percentage
            })
        
        # Сортируем по проценту совпадения
        results.sort(key=lambda x: x['match_percentage'], reverse=True)
        
        context = {
            'results': results,
            'selected_options': selected_options,
            'answered_questions': len(selected_options)
        }
        return render(request, 'identifier_results.html', context)
    
    context = {
        'characteristics': characteristics,
    }
    return render(request, 'identifier_questions.html', context)

def find_matching_mushrooms(selected_options):
    """Находит грибы по выбранным характеристикам"""
    from django.db.models import Q
    
    # Если не выбрано ни одной характеристики - показываем все грибы
    if not selected_options:
        return Mushroom.objects.all()
    
    # Начинаем с всех грибов
    mushrooms = Mushroom.objects.all()
    
    # Для каждой выбранной характеристики добавляем фильтр
    for char_id, option_value in selected_options.items():
        # Ищем грибы, у которых есть связь с этой характеристикой и значением
        mushrooms = mushrooms.filter(
            mushroomcharacteristic__characteristic_id=char_id,
            mushroomcharacteristic__option__value=option_value
        )
    
    return mushrooms.distinct()

def calculate_match_percentage(mushroom, selected_options):
    """Рассчитывает процент совпадения характеристик"""
    # Если не выбрано ни одной характеристики - 100% совпадение
    if not selected_options:
        return 100
    
    total_selected = len(selected_options)
    match_count = 0
    
    # Получаем все характеристики этого гриба
    mushroom_chars = MushroomCharacteristic.objects.filter(mushroom=mushroom)
    mushroom_char_dict = {
        mc.characteristic_id: mc.option.value 
        for mc in mushroom_chars
    }
    
    # Считаем совпадения
    for char_id, selected_value in selected_options.items():
        if char_id in mushroom_char_dict:
            if mushroom_char_dict[char_id] == selected_value:
                match_count += 1
    
    return int((match_count / total_selected) * 100) if total_selected > 0 else 100

def mushroom_detail(request, mushroom_id):
    """Детальная страница гриба с информацией о двойниках"""
    mushroom = get_object_or_404(Mushroom, id=mushroom_id)
    
    # Получаем двойников этого гриба
    lookalikes = Lookalike.objects.filter(mushroom=mushroom).select_related('lookalike')
    
    # Получаем характеристики гриба
    characteristics = MushroomCharacteristic.objects.filter(
        mushroom=mushroom
    ).select_related('characteristic', 'option')
    
    # Находим похожие грибы (по типу и съедобности)
    similar_mushrooms = Mushroom.objects.filter(
        mushroom_type=mushroom.mushroom_type,
        edibility=mushroom.edibility
    ).exclude(id=mushroom.id)[:4]
    
    context = {
        'mushroom': mushroom,
        'lookalikes': lookalikes,
        'characteristics': characteristics,
        'similar_mushrooms': similar_mushrooms,
    }
    return render(request, 'mushroom_detail.html', context)

def search_mushrooms(request):
    """Поиск грибов по названию и характеристикам"""
    query = request.GET.get('q', '')
    
    if query:
        mushrooms = Mushroom.objects.filter(
            Q(russian_name__icontains=query) |
            Q(latin_name__icontains=query) |
            Q(description__icontains=query) |
            Q(habitat__icontains=query)
        )
    else:
        mushrooms = Mushroom.objects.none()
    
    context = {
        'mushrooms': mushrooms,
        'query': query,
        'results_count': mushrooms.count()
    }
    return render(request, 'search_results.html', context)

# =============================================================================
# ПРЕДСТАВЛЕНИЯ ДЛЯ КВИЗА
# =============================================================================

def quiz_home(request):
    """Главная страница квиза с выбором уровня"""
    quizzes = Quiz.objects.all()
    
    # Показываем топ-5 результатов для каждого квиза
    quiz_results = {}
    for quiz in quizzes:
        quiz_results[quiz.id] = QuizResult.objects.filter(
            quiz=quiz
        ).order_by('-percentage', '-score')[:5]
    
    return render(request, 'quiz/quiz_home.html', {
        'quizzes': quizzes,
        'quiz_results': quiz_results
    })

def quiz_start(request, quiz_id):
    """Начало квиза - запрос имени и инициализация сессии"""
    quiz = get_object_or_404(Quiz, id=quiz_id)
    
    if request.method == 'POST':
        # Получаем имя пользователя из формы
        user_name = request.POST.get('user_name', '').strip()
        if not user_name:
            return render(request, 'quiz/quiz_start.html', {
                'quiz': quiz,
                'error': 'Пожалуйста, введите ваше имя'
            })
        
        # Получаем все вопросы для квиза
        questions = list(quiz.questions.all())
        
        if not questions:
            messages.error(request, 'В этом квизе пока нет вопросов')
            return redirect('quiz_home')
        
        # Перемешиваем вопросы
        random.shuffle(questions)
        
        # Инициализируем сессию для квиза
        request.session['quiz_data'] = {
            'quiz_id': quiz_id,
            'user_name': user_name,
            'current_question_index': 0,
            'score': 0,
            'total_questions': len(questions),
            'correct_answers': 0,
            'wrong_answers': 0,
            'question_ids': [q.id for q in questions],
            'user_answers': []  # Добавляем список для хранения ответов пользователя
        }
        
        return redirect('quiz_question')
    
    # GET запрос - показываем форму ввода имени
    return render(request, 'quiz/quiz_start.html', {'quiz': quiz})

def quiz_question(request):
    """Отображение текущего вопроса"""
    quiz_data = request.session.get('quiz_data')
    
    # Проверяем, инициализирована ли сессия квиза
    if not quiz_data:
        return redirect('quiz_home')
    
    current_index = quiz_data['current_question_index']
    question_ids = quiz_data['question_ids']
    
    # Проверяем, закончились ли вопросы
    if current_index >= len(question_ids):
        return redirect('quiz_result')
    
    # Получаем текущий вопрос
    try:
        question_id = question_ids[current_index]
        question = get_object_or_404(QuizQuestion, id=question_id)
        quiz = get_object_or_404(Quiz, id=quiz_data['quiz_id'])
        
        # Перемешиваем ответы
        answers = list(question.answers.all())
        random.shuffle(answers)
        
        # Рассчитываем прогресс
        progress = int((current_index / len(question_ids)) * 100)
        
        context = {
            'quiz': quiz,
            'question': question,
            'answers': answers,
            'current_question': current_index + 1,
            'total_questions': len(question_ids),
            'score': quiz_data['score'],
            'correct_answers': quiz_data['correct_answers'],
            'wrong_answers': quiz_data['wrong_answers'],
            'progress': progress
        }
        
        return render(request, 'quiz/quiz_question.html', context)
        
    except (IndexError, QuizQuestion.DoesNotExist, Quiz.DoesNotExist) as e:
        # Если произошла ошибка, перенаправляем на результаты
        print(f"Error in quiz_question: {e}")
        return redirect('quiz_result')

@csrf_exempt
def quiz_check_answer(request):
    """Проверка ответа (AJAX) с сохранением ответов пользователя"""
    if request.method == 'POST':
        try:
            quiz_data = request.session.get('quiz_data')
            if not quiz_data:
                return JsonResponse({'error': 'Сессия не найдена'}, status=400)
            
            current_index = quiz_data['current_question_index']
            question_ids = quiz_data['question_ids']
            
            if current_index >= len(question_ids):
                return JsonResponse({'error': 'Все вопросы пройдены'}, status=400)
            
            question_id = question_ids[current_index]
            question = get_object_or_404(QuizQuestion, id=question_id)
            
            # Инициализируем список для ответов пользователя, если его нет
            if 'user_answers' not in quiz_data:
                quiz_data['user_answers'] = []
            
            user_answer_data = {
                'question_id': question_id,
                'selected_answer_ids': [],
                'is_correct': False,
                'question_text': question.question_text,
                'question_type': question.question_type
            }
            
            # Обработка ответа в зависимости от типа вопроса
            if question.question_type == 'multiple':
                answer_ids = request.POST.getlist('answer_ids')
                if not answer_ids:
                    return JsonResponse({'error': 'Выберите хотя бы один ответ'}, status=400)
                
                selected_answers = QuizAnswer.objects.filter(id__in=answer_ids, question=question)
                correct_answers = question.answers.filter(is_correct=True)
                
                # Сохраняем ID выбранных ответов
                user_answer_data['selected_answer_ids'] = [int(aid) for aid in answer_ids]
                user_answer_data['selected_answer_texts'] = [ans.answer_text for ans in selected_answers]
                
                # Проверяем, все ли правильные ответы выбраны и нет ли неправильных
                selected_correct = selected_answers.filter(is_correct=True)
                is_correct = (selected_correct.count() == correct_answers.count() and 
                             selected_answers.count() == correct_answers.count())
                
                correct_answer_texts = [ans.answer_text for ans in correct_answers]
                correct_answer_ids = [ans.id for ans in correct_answers]
                
            else:
                answer_id = request.POST.get('answer_id')
                if not answer_id:
                    return JsonResponse({'error': 'Выберите ответ'}, status=400)
                
                selected_answer = get_object_or_404(QuizAnswer, id=answer_id, question=question)
                is_correct = selected_answer.is_correct
                correct_answers = question.answers.filter(is_correct=True)
                correct_answer_texts = [ans.answer_text for ans in correct_answers]
                correct_answer_ids = [ans.id for ans in correct_answers]
                
                # Сохраняем ID выбранного ответа
                user_answer_data['selected_answer_ids'] = [int(answer_id)]
                user_answer_data['selected_answer_texts'] = [selected_answer.answer_text]
            
            user_answer_data['is_correct'] = is_correct
            user_answer_data['correct_answer_ids'] = correct_answer_ids
            user_answer_data['correct_answer_texts'] = correct_answer_texts
            
            # Сохраняем ответ пользователя в сессии
            quiz_data['user_answers'].append(user_answer_data)
            
            # Обновляем статистику
            if is_correct:
                quiz_data['score'] += 10
                quiz_data['correct_answers'] += 1
            else:
                quiz_data['wrong_answers'] += 1
            
            # Переходим к следующему вопросу
            quiz_data['current_question_index'] += 1
            
            # Сохраняем обновленные данные в сессии
            request.session['quiz_data'] = quiz_data
            request.session.modified = True
            
            return JsonResponse({
                'is_correct': is_correct,
                'correct_answers': correct_answer_texts,
                'correct_answer_ids': correct_answer_ids,
                'score': quiz_data['score'],
                'correct_answers_count': quiz_data['correct_answers'],
                'wrong_answers_count': quiz_data['wrong_answers'],
                'next_question_index': quiz_data['current_question_index'] + 1
            })
            
        except Exception as e:
            print(f"Error in quiz_check_answer: {e}")
            return JsonResponse({'error': 'Внутренняя ошибка сервера'}, status=500)
    
    return JsonResponse({'error': 'Неверный метод запроса'}, status=400)

def quiz_result(request):
    """Отображение результатов квиза и сохранение в БД"""
    quiz_data = request.session.get('quiz_data')
    
    if not quiz_data:
        return redirect('quiz_home')
    
    quiz = get_object_or_404(Quiz, id=quiz_data['quiz_id'])
    
    total_questions = quiz_data['total_questions']
    correct_answers = quiz_data['correct_answers']
    percentage = int((correct_answers / total_questions) * 100) if total_questions > 0 else 0
    
    # Сохраняем результат в базу данных
    quiz_result_obj = QuizResult.objects.create(
        user_name=quiz_data['user_name'],
        quiz=quiz,
        score=quiz_data['score'],
        correct_answers=correct_answers,
        wrong_answers=quiz_data['wrong_answers'],
        total_questions=total_questions,
        percentage=percentage
    )
    
    # Сохраняем ответы пользователя в базу данных
    user_answers_data = quiz_data.get('user_answers', [])
    for ua_data in user_answers_data:
        question = QuizQuestion.objects.get(id=ua_data['question_id'])
        user_answer = UserAnswer.objects.create(
            quiz_result=quiz_result_obj,
            question=question,
            is_correct=ua_data['is_correct']
        )
        # Добавляем выбранные ответы
        selected_answers = QuizAnswer.objects.filter(id__in=ua_data['selected_answer_ids'])
        user_answer.selected_answers.set(selected_answers)
    
    # Получаем ID следующих квизов для рекомендаций
    quizzes = Quiz.objects.all()
    basic_quiz = quizzes.filter(level='basic').first()
    advanced_quiz = quizzes.filter(level='advanced').first()
    expert_quiz = quizzes.filter(level='expert').first()
    
    # Получаем топ-5 результатов для этого квиза
    top_results = QuizResult.objects.filter(
        quiz=quiz
    ).order_by('-percentage', '-score')[:5]
    
    # Подготавливаем детальную информацию для шаблона
    detailed_results = []
    for ua_data in user_answers_data:
        question = QuizQuestion.objects.get(id=ua_data['question_id'])
        detailed_results.append({
            'question': question,
            'user_answers': QuizAnswer.objects.filter(id__in=ua_data['selected_answer_ids']),
            'correct_answers': QuizAnswer.objects.filter(id__in=ua_data['correct_answer_ids']),
            'is_correct': ua_data['is_correct']
        })
    
    context = {
        'quiz': quiz,
        'user_name': quiz_data['user_name'],
        'score': quiz_data['score'],
        'correct_answers': correct_answers,
        'wrong_answers': quiz_data['wrong_answers'],
        'total_questions': total_questions,
        'percentage': percentage,
        'basic_quiz_id': basic_quiz.id if basic_quiz else None,
        'advanced_quiz_id': advanced_quiz.id if advanced_quiz else None,
        'expert_quiz_id': expert_quiz.id if expert_quiz else None,
        'top_results': top_results,
        'quiz_result_id': quiz_result_obj.id,
        'detailed_results': detailed_results,
        'user_answers_data': user_answers_data
    }
    
    # Очищаем сессию квиза
    if 'quiz_data' in request.session:
        del request.session['quiz_data']
    
    return render(request, 'quiz/quiz_result.html', context)

def quiz_results_detail(request, result_id):
    """Детальная страница результатов квиза"""
    quiz_result_obj = get_object_or_404(QuizResult, id=result_id)
    user_answers = UserAnswer.objects.filter(quiz_result=quiz_result_obj).select_related('question')
    
    detailed_results = []
    for ua in user_answers:
        detailed_results.append({
            'question': ua.question,
            'user_answers': ua.selected_answers.all(),
            'correct_answers': ua.question.answers.filter(is_correct=True),
            'is_correct': ua.is_correct
        })
    
    context = {
        'quiz_result': quiz_result_obj,
        'detailed_results': detailed_results,
        'total_questions': quiz_result_obj.total_questions
    }
    
    return render(request, 'quiz/quiz_results_detail.html', context)

def export_quiz_results(request):
    """Выгрузка результатов квизов в Excel"""
    # Получаем все результаты
    results = QuizResult.objects.select_related('quiz').all()
    
    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты квизов"
    
    # Заголовки
    headers = [
        'ID', 'Имя пользователя', 'Квиз', 'Уровень сложности',
        'Баллы', 'Правильные ответы', 'Неправильные ответы', 
        'Всего вопросов', 'Процент правильных', 'Дата прохождения'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Данные
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result.id)
        ws.cell(row=row, column=2, value=result.user_name)
        ws.cell(row=row, column=3, value=result.quiz.name)
        ws.cell(row=row, column=4, value=result.quiz.get_level_display())
        ws.cell(row=row, column=5, value=result.score)
        ws.cell(row=row, column=6, value=result.correct_answers)
        ws.cell(row=row, column=7, value=result.wrong_answers)
        ws.cell(row=row, column=8, value=result.total_questions)
        ws.cell(row=row, column=9, value=result.percentage)
        ws.cell(row=row, column=10, value=result.created_at.strftime('%d.%m.%Y %H:%M'))
    
    # Настройка ширины колонок
    column_widths = [8, 20, 25, 15, 8, 15, 18, 15, 18, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Создаем HTTP response с файлом Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=quiz_results_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    
    wb.save(response)
    return response

def kingdom_info(request):
    """Страница с информацией о царстве грибов"""
    return render(request, 'kingdom.html')