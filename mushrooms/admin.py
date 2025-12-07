import csv
import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import admin
from django.utils.html import format_html
from .models import Mushroom, Quiz, QuizQuestion, QuizAnswer, QuizResult, Characteristic, CharacteristicOption, MushroomCharacteristic, Lookalike, UserAnswer

class QuizAnswerInline(admin.TabularInline):
    """Встроенное редактирование ответов в вопросах"""
    model = QuizAnswer
    extra = 3
    min_num = 2
    max_num = 6

class MushroomCharacteristicInline(admin.TabularInline):
    """Встроенное редактирование характеристик грибов"""
    model = MushroomCharacteristic
    extra = 1
    autocomplete_fields = ['characteristic', 'option']

class LookalikeInline(admin.TabularInline):
    """Встроенное редактирование двойников"""
    model = Lookalike
    fk_name = 'mushroom'
    extra = 1
    autocomplete_fields = ['lookalike']

class CharacteristicOptionInline(admin.TabularInline):
    """Встроенное редактирование вариантов характеристик"""
    model = CharacteristicOption
    extra = 3

class UserAnswerInline(admin.TabularInline):
    """Встроенное отображение ответов пользователя в результатах квиза"""
    model = UserAnswer
    extra = 0
    max_num = 0
    can_delete = False
    readonly_fields = ['question', 'get_selected_answers_text', 'is_correct_display', 'get_correct_answers_text']
    fields = ['question', 'get_selected_answers_text', 'is_correct_display', 'get_correct_answers_text']
    
    def get_selected_answers_text(self, obj):
        """Текст выбранных ответов"""
        return obj.get_selected_answers_text()
    get_selected_answers_text.short_description = 'Выбранные ответы'
    
    def is_correct_display(self, obj):
        """Красивое отображение правильности ответа"""
        if obj.is_correct:
            return format_html('<span style="color: green; font-weight: bold;">✓ Правильно</span>')
        else:
            return format_html('<span style="color: red; font-weight: bold;">✗ Неправильно</span>')
    is_correct_display.short_description = 'Статус'
    
    def get_correct_answers_text(self, obj):
        """Показывает правильные ответы для неправильных ответов"""
        if not obj.is_correct:
            return format_html('<span style="color: green;">Правильные ответы: {}</span>', obj.get_correct_answers_text())
        return format_html('<span style="color: green;">—</span>')
    get_correct_answers_text.short_description = 'Правильные ответы'
    
    def has_add_permission(self, request, obj):
        return False

@admin.register(Mushroom)
class MushroomAdmin(admin.ModelAdmin):
    list_display = ['russian_name', 'latin_name', 'mushroom_type', 'edibility', 'season', 'get_lookalikes_count']
    list_filter = ['mushroom_type', 'edibility', 'season']
    search_fields = ['russian_name', 'latin_name', 'description']
    list_per_page = 20
    inlines = [MushroomCharacteristicInline, LookalikeInline]
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('russian_name', 'latin_name', 'mushroom_type', 'edibility')
        }),
        ('Описание', {
            'fields': ('description', 'habitat', 'season', 'distribution')
        }),
        ('Дополнительная информация', {
            'fields': ('key_characteristics', 'warning', 'cooking_tips'),
            'classes': ('collapse',)
        }),
        ('Фотография', {
            'fields': ('photo',),
            'classes': ('collapse',)
        }),
    )

    def get_lookalikes_count(self, obj):
        return obj.main_mushroom_lookalikes.count()
    get_lookalikes_count.short_description = 'Двойников'

@admin.register(Quiz)
class QuizAdmin(admin.ModelAdmin):
    list_display = ['name', 'level', 'questions_count', 'get_actual_questions_count', 'get_results_count']
    list_filter = ['level']
    search_fields = ['name', 'description']
    
    def get_actual_questions_count(self, obj):
        return obj.questions.count()
    get_actual_questions_count.short_description = 'Фактически вопросов'

    def get_results_count(self, obj):
        return obj.quizresult_set.count()
    get_results_count.short_description = 'Результатов'

@admin.register(QuizQuestion)
class QuizQuestionAdmin(admin.ModelAdmin):
    list_display = ['short_question_text', 'quiz', 'order', 'question_type', 'get_answers_count', 'get_correct_answers_count']
    list_filter = ['quiz', 'question_type']
    search_fields = ['question_text']
    ordering = ['quiz', 'order']
    list_per_page = 20
    inlines = [QuizAnswerInline]
    
    def short_question_text(self, obj):
        return obj.question_text[:100] + "..." if len(obj.question_text) > 100 else obj.question_text
    short_question_text.short_description = 'Текст вопроса'
    
    def get_answers_count(self, obj):
        return obj.answers.count()
    get_answers_count.short_description = 'Ответов'

    def get_correct_answers_count(self, obj):
        return obj.answers.filter(is_correct=True).count()
    get_correct_answers_count.short_description = 'Правильных'

@admin.register(QuizAnswer)
class QuizAnswerAdmin(admin.ModelAdmin):
    list_display = ['short_answer_text', 'question', 'is_correct', 'get_quiz_name']
    list_filter = ['question__quiz', 'is_correct']
    search_fields = ['answer_text', 'question__question_text']
    list_editable = ['is_correct']
    list_per_page = 50
    
    def short_answer_text(self, obj):
        return obj.answer_text[:80] + "..." if len(obj.answer_text) > 80 else obj.answer_text
    short_answer_text.short_description = 'Текст ответа'
    
    def get_quiz_name(self, obj):
        return obj.question.quiz.name
    get_quiz_name.short_description = 'Квиз'

@admin.register(QuizResult)
class QuizResultAdmin(admin.ModelAdmin):
    list_display = ['user_name', 'quiz', 'percentage', 'score', 'correct_answers', 'total_questions', 'created_at', 'get_wrong_answers_count', 'view_wrong_answers']
    list_filter = ['quiz', 'created_at']
    search_fields = ['user_name']
    readonly_fields = ['created_at']
    list_per_page = 25
    date_hierarchy = 'created_at'
    inlines = [UserAnswerInline]
    
    actions = ['export_to_excel', 'export_to_csv', 'export_detailed_results_to_excel']
    
    def get_wrong_answers_count(self, obj):
        """Количество неправильных ответов"""
        wrong_count = obj.user_answers.filter(is_correct=False).count()
        if wrong_count > 0:
            return format_html('<span style="color: red; font-weight: bold;">{}</span>', wrong_count)
        return format_html('<span style="color: green;">{}</span>', wrong_count)
    get_wrong_answers_count.short_description = 'Ошибок'
    
    def view_wrong_answers(self, obj):
        """Ссылка для просмотра неправильных ответов"""
        wrong_count = obj.user_answers.filter(is_correct=False).count()
        if wrong_count > 0:
            url = f'/admin/quiz/useranswer/?quiz_result__id__exact={obj.id}&is_correct__exact=0'
            return format_html('<a href="{}" style="background: #ff4444; color: white; padding: 5px 10px; border-radius: 3px; text-decoration: none;">Просмотреть ошибки ({})</a>', url, wrong_count)
        return format_html('<span style="color: green;">Нет ошибок</span>')
    view_wrong_answers.short_description = 'Действия'
    
    def export_to_excel(self, request, queryset):
        """Экспорт основных результатов в Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты квизов"
        
        headers = [
            'ID', 'Имя пользователя', 'Квиз', 'Уровень сложности',
            'Баллы', 'Правильные ответы', 'Неправильные ответы', 
            'Всего вопросов', 'Процент правильных', 'Дата прохождения'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
        for row, result in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=result.id)
            ws.cell(row=row, column=2, value=result.user_name)
            ws.cell(row=row, column=3, value=result.quiz.name)
            ws.cell(row=row, column=4, value=result.quiz.get_level_display())
            ws.cell(row=row, column=5, value=result.score)
            ws.cell(row=row, column=6, value=result.correct_answers)
            ws.cell(row=row, column=7, value=result.wrong_answers)
            ws.cell(row=row, column=8, value=result.total_questions)
            ws.cell(row=row, column=9, value=result.percentage)
            ws.cell(row=row, column=10, value=result.created_at.strftime('%d.%m.%Y %H:%M'))
        
        column_widths = [8, 20, 25, 15, 8, 15, 18, 15, 18, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=quiz_results_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        wb.save(response)
        return response
    
    export_to_excel.short_description = "📊 Экспорт основных результатов в Excel"
    
    def export_detailed_results_to_excel(self, request, queryset):
        """Экспорт детальных результатов с ответами пользователей"""
        wb = openpyxl.Workbook()
        
        # Лист с основными результатами
        ws_main = wb.active
        ws_main.title = "Основные результаты"
        
        main_headers = [
            'ID результата', 'Имя пользователя', 'Квиз', 'Уровень сложности',
            'Баллы', 'Правильные ответы', 'Неправильные ответы', 
            'Всего вопросов', 'Процент правильных', 'Дата прохождения'
        ]
        
        for col, header in enumerate(main_headers, 1):
            ws_main.cell(row=1, column=col, value=header)
            ws_main.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
        for row, result in enumerate(queryset, 2):
            ws_main.cell(row=row, column=1, value=result.id)
            ws_main.cell(row=row, column=2, value=result.user_name)
            ws_main.cell(row=row, column=3, value=result.quiz.name)
            ws_main.cell(row=row, column=4, value=result.quiz.get_level_display())
            ws_main.cell(row=row, column=5, value=result.score)
            ws_main.cell(row=row, column=6, value=result.correct_answers)
            ws_main.cell(row=row, column=7, value=result.wrong_answers)
            ws_main.cell(row=row, column=8, value=result.total_questions)
            ws_main.cell(row=row, column=9, value=result.percentage)
            ws_main.cell(row=row, column=10, value=result.created_at.strftime('%d.%m.%Y %H:%M'))
        
        # Настройка ширины колонок для основного листа
        main_column_widths = [12, 20, 25, 15, 8, 15, 18, 15, 18, 20]
        for i, width in enumerate(main_column_widths, 1):
            ws_main.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Лист с детальными ответами
        ws_details = wb.create_sheet(title="Детальные ответы")
        
        detail_headers = [
            'ID результата', 'Имя пользователя', 'Квиз', 'Вопрос',
            'Выбранные ответы', 'Правильные ответы', 'Статус ответа'
        ]
        
        for col, header in enumerate(detail_headers, 1):
            ws_details.cell(row=1, column=col, value=header)
            ws_details.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
        row_counter = 2
        for result in queryset:
            user_answers = result.user_answers.select_related('question').prefetch_related('selected_answers')
            
            for user_answer in user_answers:
                ws_details.cell(row=row_counter, column=1, value=result.id)
                ws_details.cell(row=row_counter, column=2, value=result.user_name)
                ws_details.cell(row=row_counter, column=3, value=result.quiz.name)
                ws_details.cell(row=row_counter, column=4, value=user_answer.question.question_text)
                ws_details.cell(row=row_counter, column=5, value=user_answer.get_selected_answers_text())
                ws_details.cell(row=row_counter, column=6, value=user_answer.get_correct_answers_text())
                ws_details.cell(row=row_counter, column=7, value="Правильно" if user_answer.is_correct else "Неправильно")
                
                row_counter += 1
        
        # Настройка ширины колонок для детального листа
        detail_column_widths = [12, 20, 25, 50, 50, 50, 15]
        for i, width in enumerate(detail_column_widths, 1):
            ws_details.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=detailed_quiz_results_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        
        wb.save(response)
        return response
    
    export_detailed_results_to_excel.short_description = "📋 Экспорт детальных результатов с ответами"
    
    def export_to_csv(self, request, queryset):
        """Экспорт выбранных результатов в CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="quiz_results_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
        response.write('\ufeff'.encode('utf8'))
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow([
            'ID', 'Имя пользователя', 'Квиз', 'Уровень сложности',
            'Баллы', 'Правильные ответы', 'Неправильные ответы', 
            'Всего вопросов', 'Процент правильных', 'Дата прохождения'
        ])
        
        for result in queryset:
            writer.writerow([
                result.id,
                result.user_name,
                result.quiz.name,
                result.quiz.get_level_display(),
                result.score,
                result.correct_answers,
                result.wrong_answers,
                result.total_questions,
                result.percentage,
                result.created_at.strftime('%d.%m.%Y %H:%M')
            ])
        
        return response
    
    export_to_csv.short_description = "📄 Экспорт выбранных результатов в CSV"
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('user_name', 'quiz', 'created_at')
        }),
        ('Результаты', {
            'fields': ('score', 'correct_answers', 'wrong_answers', 'total_questions', 'percentage')
        }),
    )

@admin.register(Characteristic)
class CharacteristicAdmin(admin.ModelAdmin):
    list_display = ['name', 'question', 'order', 'is_important', 'get_options_count']
    list_filter = ['is_important']
    search_fields = ['name', 'question']
    ordering = ['order']
    list_editable = ['order', 'is_important']
    inlines = [CharacteristicOptionInline]

    def get_options_count(self, obj):
        return obj.characteristicoption_set.count()
    get_options_count.short_description = 'Вариантов'

@admin.register(CharacteristicOption)
class CharacteristicOptionAdmin(admin.ModelAdmin):
    list_display = ['value', 'description', 'characteristic']
    list_filter = ['characteristic']
    search_fields = ['value', 'description']
    list_per_page = 30

@admin.register(MushroomCharacteristic)
class MushroomCharacteristicAdmin(admin.ModelAdmin):
    list_display = ['mushroom', 'characteristic', 'option']
    list_filter = ['characteristic', 'mushroom__mushroom_type']
    search_fields = ['mushroom__russian_name', 'characteristic__name']
    autocomplete_fields = ['mushroom', 'characteristic', 'option']
    list_per_page = 50

@admin.register(Lookalike)
class LookalikeAdmin(admin.ModelAdmin):
    list_display = ['mushroom', 'lookalike', 'danger_level', 'get_lookalike_edibility']
    list_filter = ['danger_level', 'mushroom__edibility', 'lookalike__edibility']
    search_fields = ['mushroom__russian_name', 'lookalike__russian_name']
    autocomplete_fields = ['mushroom', 'lookalike']
    
    def get_lookalike_edibility(self, obj):
        return obj.lookalike.get_edibility_display()
    get_lookalike_edibility.short_description = 'Съедобность двойника'

    fieldsets = (
        ('Основная информация', {
            'fields': ('mushroom', 'lookalike', 'danger_level')
        }),
        ('Отличия', {
            'fields': ('differences', 'visual_differences', 'warning')
        }),
    )

@admin.register(UserAnswer)
class UserAnswerAdmin(admin.ModelAdmin):
    list_display = ['quiz_result', 'short_question_text', 'get_selected_answers_display', 'is_correct_display', 'created_at', 'view_correct_answers']
    list_filter = ['is_correct', 'quiz_result__quiz', 'created_at']
    search_fields = ['quiz_result__user_name', 'question__question_text']
    readonly_fields = ['created_at', 'get_selected_answers_display', 'get_correct_answers_display']
    list_per_page = 50
    
    def short_question_text(self, obj):
        """Сокращенный текст вопроса"""
        return obj.question.question_text[:80] + "..." if len(obj.question.question_text) > 80 else obj.question.question_text
    short_question_text.short_description = 'Вопрос'
    
    def get_selected_answers_display(self, obj):
        """Отображение выбранных ответов"""
        selected_text = obj.get_selected_answers_text()
        if not obj.is_correct:
            return format_html('<span style="color: red;">{}</span>', selected_text)
        return format_html('<span style="color: green;">{}</span>', selected_text)
    get_selected_answers_display.short_description = 'Выбранные ответы'
    
    def is_correct_display(self, obj):
        """Красивое отображение правильности ответа"""
        if obj.is_correct:
            return format_html('<span style="color: green; font-weight: bold;">✓ Правильно</span>')
        else:
            return format_html('<span style="color: red; font-weight: bold;">✗ Неправильно</span>')
    is_correct_display.short_description = 'Правильность'
    
    def view_correct_answers(self, obj):
        """Показывает правильные ответы для неправильных ответов"""
        if not obj.is_correct:
            correct_text = obj.get_correct_answers_text()
            return format_html('<span style="color: green;">{}</span>', correct_text)
        return format_html('<span style="color: green;">—</span>')
    view_correct_answers.short_description = 'Правильные ответы'
    
    def get_correct_answers_display(self, obj):
        """Отображает правильные ответы в детальном просмотре"""
        correct_text = obj.get_correct_answers_text()
        return format_html(
            '<div style="background: #e8f5e8; padding: 10px; border-radius: 5px; margin: 10px 0;">'
            '<strong>Правильные ответы:</strong><br>{}'
            '</div>', 
            correct_text
        )
    get_correct_answers_display.short_description = 'Информация о правильных ответах'
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('quiz_result', 'question', 'created_at')
        }),
        ('Ответ пользователя', {
            'fields': ('get_selected_answers_display', 'is_correct')
        }),
        ('Правильные ответы', {
            'fields': ('get_correct_answers_display',)
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('quiz_result', 'question').prefetch_related('selected_answers')